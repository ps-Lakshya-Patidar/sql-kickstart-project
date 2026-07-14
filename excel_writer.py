"""
excel_writer.py
===============
Excel output layer for the dvdrental → Excel export project.

Responsibilities
----------------
- Style constants  : shared Font, Fill, and Alignment objects.
- autofit_columns(): Resize worksheet columns to fit content.
- write_sheet()    : Create a formatted worksheet for one query result.
- create_summary_sheet(): Build the front-page Summary worksheet.

This module has no knowledge of SQL or database connections — it receives
plain Python lists and writes them to openpyxl objects.
"""

from __future__ import annotations

import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import logger

# ──────────────────────────────────────────────────────────────────────────────
# STYLE CONSTANTS  –  defined once, reused across every sheet
# ──────────────────────────────────────────────────────────────────────────────
HEADER_FONT      = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL      = PatternFill("solid", fgColor="2F5496")   # deep navy-blue
ALT_ROW_FILL     = PatternFill("solid", fgColor="DCE6F1")   # light blue
HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)


# ──────────────────────────────────────────────────────────────────────────────
# COLUMN AUTO-FIT
# ──────────────────────────────────────────────────────────────────────────────

def autofit_columns(
    ws: Any,
    columns: list[str],
    rows: list[tuple[Any, ...]],
) -> None:
    """
    Resize each worksheet column to accommodate its widest cell value.

    Width is capped at 60 characters to prevent excessively wide columns
    when a text cell contains a long string.

    Parameters
    ----------
    ws : openpyxl worksheet
        The worksheet whose columns should be resized.
    columns : list[str]
        Column header names (used to seed the minimum width measurement).
    rows : list[tuple[Any, ...]]
        Data rows; each cell value is converted to ``str`` for measurement.
    """
    for col_idx, col_name in enumerate(columns, start=1):
        # Start with header width + small padding
        max_len = len(str(col_name)) + 2
        for row in rows:
            cell_value = row[col_idx - 1]
            cell_len   = len(str(cell_value)) if cell_value is not None else 0
            max_len    = max(max_len, cell_len + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 60)


# ──────────────────────────────────────────────────────────────────────────────
# SHEET WRITER
# ──────────────────────────────────────────────────────────────────────────────

def write_sheet(
    wb: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[tuple[Any, ...]],
) -> None:
    """
    Create and populate a named worksheet inside the workbook.

    Formatting applied
    ------------------
    - Row 1: bold, white text on navy-blue background, centred, height = 20.
    - Row 1: frozen so headers stay visible while scrolling.
    - Even data rows: light-blue alternating fill.
    - All columns: auto-fitted width (capped at 60).

    Parameters
    ----------
    wb : openpyxl.Workbook
        The target workbook.  A new sheet is appended.
    sheet_name : str
        Worksheet title (e.g. ``"Q1"``).
    columns : list[str]
        Column headers derived from ``cursor.description``.
    rows : list[tuple[Any, ...]]
        Data rows to write beneath the headers.
    """
    ws = wb.create_sheet(title=sheet_name)

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(columns, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    ws.row_dimensions[1].height = 20
    ws.freeze_panes             = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill

    # ── Auto-fit columns ──────────────────────────────────────────────────────
    autofit_columns(ws, columns, rows)
    logger.debug(
        "[%s]  Sheet written  (%d cols × %d rows)",
        sheet_name, len(columns), len(rows),
    )


# ──────────────────────────────────────────────────────────────────────────────
# SUMMARY SHEET
# ──────────────────────────────────────────────────────────────────────────────

def create_summary_sheet(
    wb: Workbook,
    summary_data: list[dict[str, Any]],
) -> None:
    """
    Build the front-page "Summary" worksheet with execution metadata.

    Columns written
    ---------------
    - Question           : query label (e.g. "Q1")
    - Rows Returned      : integer row count
    - Execution Time (s) : rounded to 4 decimal places
    - Status             : "SUCCESS" (green) | "FAILED" (red) | "SKIPPED" (orange)

    The sheet is inserted at index 0 so it always appears as the first tab.

    Parameters
    ----------
    wb : openpyxl.Workbook
        The target workbook.
    summary_data : list[dict[str, Any]]
        One dict per query containing the keys:
        ``question``, ``rows_returned``, ``execution_time``, ``status``.
    """
    ws = wb.create_sheet(title="Summary", index=0)

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Question", "Rows Returned", "Execution Time (s)", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    ws.row_dimensions[1].height = 20
    ws.freeze_panes             = "A2"

    # ── Status colour map ─────────────────────────────────────────────────────
    status_colors: dict[str, str] = {
        "SUCCESS": "00B050",   # green
        "FAILED":  "FF0000",   # red
        "SKIPPED": "FFA500",   # orange
    }

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, entry in enumerate(summary_data, start=2):
        status = entry.get("status", "UNKNOWN")
        values = [
            entry["question"],
            entry["rows_returned"],
            round(entry["execution_time"], 4),
            status,
        ]
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill

        # Colour-code the Status cell independently of the row fill
        status_cell      = ws.cell(row=row_idx, column=4)
        status_cell.font = Font(
            bold=True,
            color=status_colors.get(status, "000000"),
        )

    # ── Fixed column widths (content is predictable) ──────────────────────────
    for col_idx, width in enumerate([14, 16, 22, 12], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    logger.info("Summary sheet created  (%d entries).", len(summary_data))
